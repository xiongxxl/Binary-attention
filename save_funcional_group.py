from openpyxl import load_workbook
# 指定已有的Excel文件路径


def save_functional_excel(excel_path,data_to_append):

    # 加载已有的Excel工作簿
    wb = load_workbook(filename=excel_path)
    # 选择要写入数据的工作表，这里假设是第一个工作表
    sheet = wb.active
    # 假设有一个列表，包含要写入的多行数据

    # 确定写入数据的起始行号，这里假设在现有数据之后写入
    start_row = sheet.max_row + 1
    # 逐行写入数据
    for row_data in data_to_append:
        sheet.append(row_data)
    # 保存Excel文件，覆盖原有文件
    wb.save(filename=excel_path)
    return excel_path

if __name__ == "__main__":
    excel_path = 'example.xlsx'
    data_to_append = [
        ["姓名", "年龄", "性别"],
        ["赵六", 25, "男"],
        ["孙七", 35, "女"]
    ]
    excel_path=save_functional_excel(excel_path,data_to_append)